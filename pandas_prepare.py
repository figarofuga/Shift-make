
#%%
import numpy as np
import pandas as pd
import datetime
import re
import pickle
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
#%% 
month = 8
ok_sign = r'◯|〇|◯１|希望日'
ng_sign = r'×|✕|不可日'

koureisha = [
    "上野", 
    "安富大祐", 
    "鄭東孝",
    "福島龍貴", 
    "本田美和子",
    "長谷川華子",
    "門松賢", 
    "松浦",  
]

c_member = [
    "鈴木勝也","清水隆之","小林佐紀子","里見良輔","岩瀬恭子","片山充哉","入佐薫","藤村慶子","籠尾壽哉","山田康博"
]

tochoku_member = ["福原誠一郎","太良史郎","林智史","持丸貴生","大重達寛","雪野満","吉田心慈","久冨木原健二","松永崇宏","新美望","篠﨑太郎","青木康浩","脇坂悠介","茅島敦人","津山頌章","川瀬咲","門松賢",
                  "東條誠也","莇舜平","荒金直美","佐川偲","藤澤まり",
                  "勝俣敬寛","松井貴裕","伊藤国秋","平井智大",
                #   "中瀬大","山本晴二郎","宗大輔","伊藤瑛基","中枝建郎","花岡孝行","安藤昂志",
                  "熊木聡美","織部峻太郎","臼坂優希","深沢夏海","松原龍輔","佐久川佳怜","吉田博道","佐久間一也","野上創生","髙木菜々美","谷岡友則","細川善弘","小宮健太郎","高梨航輔","福井梓穂","石井真央","岩崎文美","岡村真伊","黒崎颯","鈴木徹志郎","吉村梨沙","先﨑光","星貴文"
]

#%%
# read excel data
dat_proto = pd.read_excel(f"rawdata/{month}m/2024年{month}月の日当直・ICU勤務・1次救急希望調査（回答）.xlsx")


dat = (dat_proto
       .assign(name = lambda x: x['お名前'].str.replace('[　 ]', '', regex=True))
       .assign(doctoryear = lambda x: x['あなたは医師何年目ですか？'].replace(r'年.*', '', regex=True).astype(int))
)

#%%

comment_dat = (dat.filter(['タイムスタンプ', 'name', '日直・当直希望についての備考', '1次救急希望についての備考',
       'ICU勤務希望についての備考', 'このアンケート対する意見があればお願いします。', 'doctoryear'])
       .assign(タイムスタンプ=lambda x: pd.to_datetime(x['タイムスタンプ']))
                     .sort_values(by='タイムスタンプ', ascending=True)
                     .groupby('name')
                     .last()
                     .drop(['タイムスタンプ'], axis=1)
                     .reset_index()
)



# 列名を変更する関数
def extract_date(column_name):
    match = re.search(r'\d+月\d+日', column_name)
    if match:
        return match.group()  # 日付部分のみを返す
    return column_name  # マッチしない場合は元の列名を返す

# base data
base_data = (dat
             .filter(['タイムスタンプ', 'name', 'あなたの専門はなんですか?', 'あなたは医師何年目ですか?'])
             )
#%%

kiboubi_df = (dat.filter(regex=r'タイムスタンプ|name|日直・当直希望.*\d{1,2}月')
       .rename(columns=extract_date)
       .melt(id_vars=['タイムスタンプ', 'name'], 
             var_name='date', value_name='request')
       .dropna(axis=0, subset=['request'])
       .query('request=="希望日"')
       .groupby('date')['name']
       .agg(' ,'.join)
       .reset_index()
       .assign(date = lambda x: pd.to_datetime('2024年' + x['date'], format='%Y年%m月%d日'))
       .sort_values(by='date', ascending=True)
       .reset_index(drop=True)

)

fukabi_df = (dat.filter(regex=r'タイムスタンプ|name|日直・当直希望.*\d{1,2}月')
       .rename(columns=extract_date)
       .melt(id_vars=['タイムスタンプ', 'name'], 
             var_name='date', value_name='request')
       .dropna(axis=0, subset=['request'])
       .query('request=="不可日"')
       .groupby('date')['name']
       .agg(' ,'.join)
       .reset_index()
       .assign(date = lambda x: pd.to_datetime('2024年' + x['date'], format='%Y年%m月%d日'))
       .sort_values(by='date', ascending=True)
       .reset_index(drop=True)
)

#%%
# C日直のデータを整形
data_c = (dat
        .filter(regex=r'日直・当直希望.*\d{1,2}月')
        .rename(columns=extract_date)
)
combined_data_c = (pd.concat([base_data, data_c], axis=1)
                       .query('name in @c_member')
                       .assign(タイムスタンプ=lambda x: pd.to_datetime(x['タイムスタンプ']))
                          .sort_values(by='タイムスタンプ', ascending=True)
                            .groupby('name')
                            .last()
                            .drop(['タイムスタンプ'], axis=1)
                            .reset_index()
                            .melt(id_vars=['name'], 
                                  var_name='date', 
                                  value_name='request', 
                                  ignore_index=False)
                            .applymap(lambda x: x.strip() if isinstance(x, str) else x)   
                            .assign(request=lambda x: np.where(x['request'] == '', None, x['request']))   
                            
)

data_wide_c_pre = (combined_data_c
             .groupby(['name', 'request'])['date']
             .apply(lambda x: ' ,'.join(x))
             .reset_index()
             .pivot(index='name', columns='request', values='date')
             .reset_index()
    )
columns_ok = data_wide_c_pre.filter(regex=ok_sign).columns
if not columns_ok.empty:
    # 該当する列のデータをコンマ区切りで結合
    data_wide_c_pre['accept'] = data_wide_c_pre[columns_ok].apply(lambda x: ', '.join(x.dropna().astype(str)), axis=1)
    
columns_bad = data_wide_c_pre.filter(regex=ng_sign).columns
if not columns_bad.empty:
    # 該当する列のデータをコンマ区切りで結合
    data_wide_c_pre['reject'] = data_wide_c_pre[columns_bad].apply(lambda x: ', '.join(x.dropna().astype(str)), axis=1)

data_wide_c = (data_wide_c_pre
                 .drop(columns=columns_ok)
                 .drop(columns=columns_bad)
                 .assign(species = 'c')
    )

comment_c = comment_dat.filter(regex=r'name|日直・当直')

data_wide_c_comment = (data_wide_c
                 .merge(comment_c, on='name', how='left')
            )


#%%
# 当直のデータを整形
data_tochoku = (dat
        .filter(regex=r'日直・当直希望.*\d{1,2}月')
        .rename(columns=extract_date)
)
combined_data_tochoku = (pd.concat([base_data, data_tochoku], axis=1)
                       .query('name in @tochoku_member')
                       .assign(タイムスタンプ=lambda x: pd.to_datetime(x['タイムスタンプ']))
                          .sort_values(by='タイムスタンプ', ascending=True)
                            .groupby('name')
                            .last()
                            .drop(['タイムスタンプ'], axis=1)
                            .reset_index()
                            .melt(id_vars=['name'], 
                                  var_name='date', 
                                  value_name='request', 
                                  ignore_index=False)
                            .applymap(lambda x: x.strip() if isinstance(x, str) else x)   
                            .assign(request=lambda x: np.where(x['request'] == '', None, x['request']))   
                            .assign(name = lambda x: x['name'].str.replace('[　 ]', '', regex=True))
                            
)

data_wide_tochoku_pre = (combined_data_tochoku
             .groupby(['name', 'request'])['date']
             .apply(lambda x: ' ,'.join(x))
             .reset_index()
             .pivot(index='name', columns='request', values='date')
             .reset_index()
    )
columns_ok = data_wide_tochoku_pre.filter(regex=ok_sign).columns
if not columns_ok.empty:
    # 該当する列のデータをコンマ区切りで結合
    data_wide_tochoku_pre['accept'] = data_wide_tochoku_pre[columns_ok].apply(lambda x: ', '.join(x.dropna().astype(str)), axis=1)
    
columns_bad = data_wide_tochoku_pre.filter(regex=ng_sign).columns
if not columns_bad.empty:
    # 該当する列のデータをコンマ区切りで結合
    data_wide_tochoku_pre['reject'] = data_wide_tochoku_pre[columns_bad].apply(lambda x: ', '.join(x.dropna().astype(str)), axis=1)

data_wide_tochoku = (data_wide_tochoku_pre
                 .drop(columns=columns_ok)
                 .drop(columns=columns_bad)
                 .assign(species = 'tochoku')
    )

comment_tochoku = comment_dat.filter(regex=r'name|日直・当直|doctoryear')

data_wide_tochoku_comment = (data_wide_tochoku
                 .merge(comment_tochoku, on='name', how='left')
                 .sort_values(by='doctoryear', ascending=True)
                 .drop(columns=['doctoryear'])
            )

#%%
# ICU勤務のデータを整形
data_icu = (dat
        .filter(regex=r'ICU勤務.*\d{1,2}月')
        .rename(columns=extract_date)
)

combined_data_icu = (pd.concat([base_data, data_icu], axis=1)
                       .assign(タイムスタンプ=lambda x: pd.to_datetime(x['タイムスタンプ']))
                          .sort_values(by='タイムスタンプ', ascending=True)
                            .groupby('name')
                            .last()
                            .drop(['タイムスタンプ'], axis=1)
                            .reset_index()
                            .melt(id_vars=['name'], 
                                  var_name='date', 
                                  value_name='request', 
                                  ignore_index=False)
                            .applymap(lambda x: x.strip() if isinstance(x, str) else x)   
                            .assign(request=lambda x: np.where(x['request'] == '', None, x['request']))   
                            .assign(name = lambda x: x['name'].str.replace('[　 ]', '', regex=True))
                           
)


data_wide_icu_pre = (combined_data_icu
             .groupby(['name', 'request'])['date']
             .apply(lambda x: ' ,'.join(x))
             .reset_index()
             .pivot(index='name', columns='request', values='date')  
             .reset_index()  
    )
columns_ok = data_wide_icu_pre.filter(regex=ok_sign).columns
if not columns_ok.empty:
    # 該当する列のデータをコンマ区切りで結合
    data_wide_icu_pre['accept'] = data_wide_icu_pre[columns_ok].apply(lambda x: ', '.join(x.dropna().astype(str)), axis=1)
    
columns_bad = data_wide_icu_pre.filter(regex=ng_sign).columns
if not columns_bad.empty:
    # 該当する列のデータをコンマ区切りで結合
    data_wide_icu_pre['reject'] = data_wide_icu_pre[columns_bad].apply(lambda x: ', '.join(x.dropna().astype(str)), axis=1)

data_wide_icu = (data_wide_icu_pre
                 .drop(columns=columns_ok)
                 .drop(columns=columns_bad)
                 .assign(species = 'icu')
    )

comment_icu = comment_dat.filter(regex=r'name|ICU勤務|doctoryear')

data_wide_icu_comment = (data_wide_icu
                 .merge(comment_icu, on='name', how='left')
                 .sort_values(by='doctoryear', ascending=True)
                 .drop(columns=['doctoryear'])
            )
#%%
# 1次救急のデータを整形
data_ichijikyu = (dat
        .filter(regex=r'1次救急.*\d{1,2}月')
        .rename(columns=extract_date)
)
data_concat_ichijikyu = (pd.concat([base_data, data_ichijikyu], axis=1)
                       .assign(タイムスタンプ=lambda x: pd.to_datetime(x['タイムスタンプ']))
                          .sort_values(by='タイムスタンプ', ascending=True)
                            .groupby('name')
                            .last()
                            .drop(['タイムスタンプ'], axis=1)
                            .reset_index()
                            .melt(id_vars=['name'], 
                                  var_name='date', 
                                  value_name='request', 
                                  ignore_index=False)
)
# notes dataと結合

combined_data_ichijikyu = (data_concat_ichijikyu
                            .applymap(lambda x: x.strip() if isinstance(x, str) else x)   
                            .assign(request=lambda x: np.where(x['request'] == '', None, x['request']))   
                            .assign(name = lambda x: x['name'].str.replace('[　 ]', '', regex=True))
                            )

data_wide_ichijikyu_pre = (combined_data_ichijikyu
             .groupby(['name', 'request'])['date']
             .apply(lambda x: ' ,'.join(x))
             .reset_index()
             .pivot(index='name', columns='request', values='date')
             .reset_index()
    )
columns_ok = data_wide_ichijikyu_pre.filter(regex=ok_sign).columns
if not columns_ok.empty:
    # 該当する列のデータをコンマ区切りで結合
    data_wide_ichijikyu_pre['accept'] = data_wide_ichijikyu_pre[columns_ok].apply(lambda x: ', '.join(x.dropna().astype(str)), axis=1)
    
columns_bad = data_wide_ichijikyu_pre.filter(regex=ng_sign).columns
if not columns_bad.empty:
    # 該当する列のデータをコンマ区切りで結合
    data_wide_ichijikyu_pre['reject'] = data_wide_ichijikyu_pre[columns_bad].apply(lambda x: ', '.join(x.dropna().astype(str)), axis=1)

data_wide_ichijikyu = (data_wide_ichijikyu_pre
                 .drop(columns=columns_ok)
                 .drop(columns=columns_bad)
                 .assign(species = 'ichijikyu')
    )

comment_ichijikyu = comment_dat.filter(regex=r'name|1次救急|doctoryear')

data_wide_ichijikyu_comment = (data_wide_ichijikyu
                 .merge(comment_ichijikyu, on='name', how='left')
                 .sort_values(by='doctoryear', ascending=True)
                 .drop(columns=['doctoryear'])
            )

# %%

kiboubi_df.to_excel(f"prepdata/{month}m/kiboubi_{month}.xlsx")
fukabi_df.to_excel(f"prepdata/{month}m/fukabi_{month}.xlsx")

#%%
# Excelファイルへのエクスポート
tochoku_file_path = f'prepdata/{month}m/tochoku_comment_{month}.xlsx'

with pd.ExcelWriter(tochoku_file_path, engine='openpyxl') as writer:
    data_wide_tochoku_comment.to_excel(writer, index=False, sheet_name='Sheet1')
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']
    
    # 列の幅を設定（2番目と3番目の列を390ピクセルに）
    worksheet.column_dimensions['B'].width = 50.86  # Excelの幅は約1/4ピクセル
    worksheet.column_dimensions['C'].width = 50.86
    worksheet.column_dimensions['E'].width = 50.86
    
    # テキストの折り返しと縦方向の自動調整
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=2, max_col=3):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    # 行の高さを自動調整
    def adjust_row_height(ws):
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=3):
            max_height = 15  # デフォルトの高さ
            for cell in row:
                if cell.value:
                    lines = cell.value.count('\n') + 1
                    # セル内の改行と列幅に基づいて行の高さを設定
                    # 大まかな計算では、列の幅ごとに行の高さを設定
                    lines += len(cell.value) // 50.86 + 1 # 列の幅に基づいて行数を計算
                    max_height = max(max_height, lines * 15)  # 1行あたり約15ピクセルで計算
            ws.row_dimensions[row[0].row].height = max_height

    adjust_row_height(worksheet)


#%%
# Excelファイルへのエクスポート
icu_file_path = f'prepdata/{month}m/icu_comment_{month}.xlsx'

with pd.ExcelWriter(icu_file_path, engine='openpyxl') as writer:
    data_wide_icu_comment.to_excel(writer, index=False, sheet_name='Sheet1')
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']
    
    # 列の幅を設定（2番目と3番目の列を390ピクセルに）
    worksheet.column_dimensions['B'].width = 50.86  # Excelの幅は約1/4ピクセル
    worksheet.column_dimensions['C'].width = 50.86
    worksheet.column_dimensions['E'].width = 50.86
    
    # テキストの折り返しと縦方向の自動調整
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=2, max_col=3):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    # 行の高さを自動調整
    def adjust_row_height(ws):
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=3):
            max_height = 15  # デフォルトの高さ
            for cell in row:
                if cell.value:
                    lines = cell.value.count('\n') + 1
                    # セル内の改行と列幅に基づいて行の高さを設定
                    # 大まかな計算では、列の幅ごとに行の高さを設定
                    lines += len(cell.value) // 50.86 + 1 # 列の幅に基づいて行数を計算
                    max_height = max(max_height, lines * 15)  # 1行あたり約15ピクセルで計算
            ws.row_dimensions[row[0].row].height = max_height

    adjust_row_height(worksheet)

#%%
# Excelファイルへのエクスポート
ichijikyu_file_path = f'prepdata/{month}m/ichijikyu_comment_{month}.xlsx'

with pd.ExcelWriter(ichijikyu_file_path, engine='openpyxl') as writer:
    data_wide_ichijikyu_comment.to_excel(writer, index=False, sheet_name='Sheet1')
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']
    
    # 列の幅を設定（2番目と3番目の列を390ピクセルに）
    worksheet.column_dimensions['B'].width = 50.86  # Excelの幅は約1/4ピクセル
    worksheet.column_dimensions['C'].width = 50.86
    worksheet.column_dimensions['E'].width = 50.86
    
    # テキストの折り返しと縦方向の自動調整
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=2, max_col=3):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    # 行の高さを自動調整
    def adjust_row_height(ws):
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=3):
            max_height = 15  # デフォルトの高さ
            for cell in row:
                if cell.value:
                    lines = cell.value.count('\n') + 1
                    # セル内の改行と列幅に基づいて行の高さを設定
                    # 大まかな計算では、列の幅ごとに行の高さを設定
                    lines += len(cell.value) // 50.86 + 1 # 列の幅に基づいて行数を計算
                    max_height = max(max_height, lines * 15)  # 1行あたり約15ピクセルで計算
            ws.row_dimensions[row[0].row].height = max_height

    adjust_row_height(worksheet)


# %%
long_dict = {
    'tochoku': combined_data_tochoku,
    'icu': combined_data_icu,
    'ichijikyu': combined_data_ichijikyu
}
  
with open(f"prepdata/{month}m/long_dict_{month}.pkl", 'wb') as f:
    pickle.dump(long_dict, f, protocol=pickle.HIGHEST_PROTOCOL)
# %%
